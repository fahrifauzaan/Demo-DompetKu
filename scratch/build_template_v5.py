#!/usr/bin/env python3
"""Template v4 → v5: tambah kolom skema baru (classification, documentLink) tanpa mengubah data contoh."""
import openpyxl
SRC, OUT = 'dompetku_template_v4.xlsx', 'dompetku_template_v5.xlsx'
ADD = {'BudgetCategories': ['classification'], 'AssetsNonLiquid': ['documentLink'], 'FixedIncome': ['documentLink']}
wb = openpyxl.load_workbook(SRC)
for tab, cols in ADD.items():
    if tab not in wb.sheetnames:
        print(f'  ! tab {tab} tak ada, dilewati'); continue
    ws = wb[tab]
    existing = [str(ws.cell(1, c).value or '').strip().lower() for c in range(1, ws.max_column + 1)]
    for col in cols:
        if col.lower() in existing:
            print(f'  – {tab}.{col} sudah ada'); continue
        idx = ws.max_column + 1
        src = ws.cell(1, ws.max_column)          # tiru gaya header tetangga
        cell = ws.cell(1, idx, col)
        cell.font, cell.fill, cell.alignment = src.font.copy(), src.fill.copy(), src.alignment.copy()
        print(f'  + {tab}.{col} @ kolom {idx}')
wb.save(OUT)
print(f'\n✅ {OUT}')
for tab in ADD:
    if tab in wb.sheetnames:
        ws = wb[tab]; print(f'   {tab}:', [ws.cell(1,c).value for c in range(1, ws.max_column+1)][-3:])
